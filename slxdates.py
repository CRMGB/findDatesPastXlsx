import os
import openpyxl
from typing import List
from datetime import date


TODAY = date.today()

class XlsxFilesDatesChecker:
    def get_files(self) -> List[str]:
        """ Get the files .xlsx from current directory."""
        current_dir = os.path.dirname(os.path.abspath(__file__)) + "/files/"
        # Find file with extension
        fileExt = r".xlsx"
        return [
            os.path.join(current_dir, _) for _ in os.listdir(current_dir) if _.endswith(fileExt)
        ]

    def load_xlsx_file_and_find_dates(self) -> None:
        """ Loop over all files from current directory
        and call method find_dates_in_past() to return the result."""
        files = self.get_files()
        for file in files:
            print(f"Loading file: {file}")
            dataframe = openpyxl.load_workbook(file)
            dat_active = dataframe.active
            result = self.find_dates_in_past(dat_active)
            if len(result)<1:
                result.append("No dates found in the PAST")
            self.edit_file(dat_active, result)
            dataframe.save(file)
        dataframe.close()
        
    def find_dates_in_past(self, dat_active) -> List[str]:
        """ Loop over all rows/columns to find the actual dates in the past
        and return the result."""    
        result = [] # type: List[str]
        for row in range(0, dat_active.max_row):
            for col in dat_active.iter_cols(1, dat_active.max_column):
                if col[row].value == None or type(col[row].value) == str:
                    continue
                if col[row].value.date()<TODAY:
                    result.append(
                        f"ALERT! Date: {col[row].value.strftime('%d/%m/%Y')} is in the PAST"
                    )
        return result

    def edit_file(self, dat_active, result) -> None:
        """ Submmit the result to the cell A1."""       
        today = TODAY.strftime('%d/%m/%Y')
        print(f"Today is: {today}")
        res = '. '.join([str(elem) for elem in result])
        print(f"RESULT---> {res}")
        dat_active.cell(row=1, column=1).value = res

init = XlsxFilesDatesChecker()
init.load_xlsx_file_and_find_dates()